from re import search
import PySimpleGUI as sg
import configparser
from pathlib import Path
import openpyxl
import pyqrcode
import gui


def main():
    # 最初に表示するウィンドウを指定する
    window = gui.main_window()

    try:
        # 設定ファイル読み込み
        config = configparser.ConfigParser()
        config.read('config.ini')
        # 初期設定が完了しているか確認
        qrfolder_path = config['setting']['qrfolder_path']
        itemlist_path = config['setting']['itemlist_path']
        # 新規登録、備品管理ボタンを有効化
        window['-Move_change_window-'].update(disabled=False)
        window['-Move_registration_window-'].update(disabled=False)
    except Exception as e:
        sg.popup('初期設定を行なってください。', e)

    # 各種イベント処理
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit' or event == '終了':
            break
        #-----
        # メイン画面のイベント
        #-----
        # 初期設定画面へ移動
        if event == '初期設定':
            window.close()
            window = gui.initialize_window()

        # 新規登録画面へ移動
        if event == '-Move_registration_window-':
            window.close()
            window = gui.registration_window()
        # 備品管理画面に移動
        if event == '-Move_change_window-':
            window.close()
            window = gui.change_window()
            # 備品一覧.xlsxから読み込み管理番号リストボックスに反映
            pass
        # 備品管理画面へ移動

        #-----
        # 初期設定画面のイベント
        #-----
        # データ保存先が選択された場合、初期設定ボタンを有効化
        if event == '-Folder_path-':
            window['-Initialize-'].update(disabled=False)

        # アプリの初期設定(必要なフォルダ、ファイルの類の作成と設定ファイル更新)
        if event == '-Initialize-':
            # データの保存先
            folder_path = values['-Folder_path-']
            # アプリの初期設定
            initialize_app(folder_path)
            # 初期設定画面をクローズし、メイン画面へ切り替え
            window.close()
            window = gui.main_window()
            # 新規登録、備品管理ボタンを有効化
            window['-Move_change_window-'].update(disabled=False)
            window['-Move_registration_window-'].update(disabled=False)

            
            
            

        #-----
        # 新規登録画面のイベント
        #-----
        # 備品の新規登録
        if event == '-Register-':
            try:
                # ウィジェットの値を取得
                input_data = [values['-Item_id-'], values['-Item_name-'],
                        values['-Shelf_number-'], values['-Owner-']]
                # 管理番号がない場合はValueErrorを発生
                if input_data[0] == '':
                    raise ValueError('管理番号を入力してください。')
                # 管理番号に重複が無いかチェック
                if is_unique_id(itemlist_path, input_data[0]) is not True:
                    raise ValueError('管理番号が重複しています。別の番号を入力下さい。')
                # 備品一覧.xlsxと備品シートを読み込み
                wb = openpyxl.load_workbook(itemlist_path)
                ws = wb['備品シート']
                # 備品一覧.xlsxのデータを変数xlsxdataに格納
                last_address = 'E' + str(ws.max_row)
                xlsxdata = ws['A2:{}'.format(last_address)]
                # 繰り返し検索
                row = 0
                while row < len(xlsxdata):
                    # 備品一覧.xlsxから一致する管理番号を検索
                    if input_data[0] == str(xlsxdata[row][0].value):
                        # 一致する行番号は、row + 2
                        matched_data = row + 2
                        break
                    else:
                        row += 1
                else:
                    matched_row = None
                
                if matched_row is not None:
                    raise ValueError('管理番号が重複しています。別の番号を入力下さい。')
                # QR code 作成
                qrcode_path = create_qrcode(qrfolder_path, input_data)
                # 備品一覧.xlsxに備品データを登録
                create_data(itemlist_path, input_data, qrcode_path)
            except Exception as e:
                sg.popup(e)
        #-----
        # 備品管理画面のイベント
        #-----

        # 備品情報の更新

        # 備品情報の削除
    
    window.close()

def initialize_app(path):
    # ディレクトリの作成
    folder_path = path
    # 備品フォルダの作成
    root_path = folder_path + '/備品管理'
    Path(root_path).mkdir()
    # QRコードフォルダの作成
    qrfolder_path = root_path + '/QRコード'
    Path(qrfolder_path).mkdir()
    # 備品一覧.xlsxを作成
    itemlist_path = root_path + '/備品一覧.xlsx'
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '備品シート'
    # 列幅の設定
    column_widths = {'A': 16, 'B': 16, 'C': 16, 'D': 16, 'E': 30}
    # 列の見出し
    column_name = {
        'A': '管理番号', 'B': '品名', 'C': '棚番号', 'D': '管理者', 'E':'QRコード'
    }
    for count, column_index in enumerate(column_widths):
        count += 1
        sheet.column_dimensions[column_index].width = column_widths[column_index]
        sheet.cell(1, count).value = column_name[column_index]
    wb.save(itemlist_path)
    # 設定ファイルのconfig.iniを更新
    config = configparser.ConfigParser()
    config.read('config.ini', 'utf-8')
    # 設定ファイルにsettingセクションがない場合のみセクション追加
    if config.has_section('setting') is not True:
        config.add_section('setting')
        config['setting']['root_path'] = root_path
        config['setting']['qrfolder_path'] = qrfolder_path
        config['setting']['itemlist_path'] = itemlist_path
        configpath = Path('config.ini')
    with configpath.open(mode='w', encoding='utf-8') as f:
        config.write(f)
    
    return sg.popup('初期設定が完了しました。')

def is_unique_id(path, search_value):
    # 備品一覧.xlsxのデータを取得
    wb, ws, xlsxdata = get_xlsx_data(path)
    # 備品一覧.xlsxに重複する管理番号があるかチェック
    if search_row(xlsxdata, search_value) is None:
        # 重複がなければ、True(重複なし)を返す
        return True

def get_xlsx_data(path):
    # 備品一覧.xlsxと備品シートを読み込み
    wb = openpyxl.load_workbook(path)
    ws = wb['備品シート']
    # 備品一覧.xlsxのデータを変数dataに格納
    last_address = 'E' + str(ws.max_row)
    data = ws['A2:{}'.format(last_address)]
    return wb, ws, data

def search_row(data, search_value):
    # 繰り返し検索
    row = 0
    while row < len(data):
        # 備品一覧.xlsxから一致する管理番号を検索
        if str(search_value) == str(data[row][0].value):
            # 一致する行番号は、row + 2
            matched_row = row + 2
            return matched_row
        else:
            row += 1
    else:
        # 一致しない場合は、Noneを返す
        return None

def create_qrcode(path, data):
    # QRコードに記録する情報を設定
    qrcontents = str(data[0]) + ';' + str(data[1]) + ';' + str(data[2]) + ';' + str(data[3])
    # QRコードを作成
    qr = pyqrcode.create(content=qrcontents, encoding='utf-8')
    # QRコードを保存するパスを設定
    qrcode_path = path + '/' + str(data[0]) + '.png'
    # QRコードをpng形式で保存
    qr.png(file=qrcode_path)

    return qrcode_path

def create_data(path, data, qr_path):
    # 備品一覧.xlsxと備品シートを読み込み
    wb = openpyxl.load_workbook(path)
    ws = wb['備品シート']
    # 追記する行番号を設定
    last_row = ws.max_row + 1
    # セルにデータを書き込み
    for colm in range(len(data)):
        ws.cell(row=last_row, column=colm + 1, value=data[colm])
    # セルにQRコードのパスを書き込み
    ws.cell(row=last_row, column=len(data) + 1, value=qr_path)
    # wbを上書き保存
    wb.save(path)
    return sg.popup_ok('書き込みが完了しました')
if __name__ == '__main__':
    main()
